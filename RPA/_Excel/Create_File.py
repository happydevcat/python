from openpyxl import Workbook

wb = Workbook() # New 워크북 생성
#ws = wb.active # 현재 활성화된 Sheet 활성화
ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성 
ws.title ="Second Sheet"
ws.sheet_properties.tabColor ="ff0066" # RGB 형태 

ws2St = wb.create_sheet("Third Sheet")
ws3St = wb.create_sheet("Fourth Sheet",0)

wsDict  = wb["Second Sheet"] # Dict 형태로 Sheet에 접근

print(wb.sheetnames)

# Sheet Copy
ws["A1"] = "뭘보냐"
cpTargetSt  = wb.copy_worksheet(ws)
cpTargetSt.title ="Copied Sheet"


wb.save("_Sample_RPA.xlsx")
wb.close()


print("사피엔스 - 유발 하라리")
print("코스모스 - 칼 세이건")
print("총, 균 쇠 - 재레드 다이아몬드")

