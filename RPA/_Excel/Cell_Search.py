from openpyxl import Workbook
from openpyxl import load_workbook # 파일 불러오기
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = load_workbook("_Sample_RPA.xlsx")
ws = wb.active # 현재 활성화된 Sheet 활성화

for row in ws.iter_rows(min_row=2) :
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어 천재")


# for row in ws.iter_rows(max_row=1) :
#     for cell in row:
#         if cell.value =="영어" :
#             cell.value ="컴퓨터"

#wb.save("_Sample_RPA_Modifyied.xlsx")

# column_B = ws["B"] # 영어 컬럼만 가지고옵니다. 
# for BCell in column_B:
#     print(BCell.value)

# column_Range = ws["B:C"] # 영어수학 가지고 오기
# for cols in column_Range : 
#     for cell in cols:
#         print(cell.value)


# # Header Info 정보 가지고 오기 
# row_title = ws[1]
# for cell in row_title:
#     print(cell.value)        

# row_range = ws[2:9] # 2번째 Row에서 9번째 Row가지고오기,. 
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()


# row_range = ws[2:ws.max_row]
# for rows in row_range:
#     for cell in rows:
#         #print(cell.value, end =" ")
#         #print(cell.coordinate, end =" ") # 좌표정보
#         xy = coordinate_from_string(cell.coordinate)
#         #print(xy, end="")
#         print(xy[0], end=" ")
#     print()

# 대상으로 row 단위로 Tuple 로 Wrap
# print(tuple(ws.rows))
# for row in tuple(ws.rows):    # for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3)
#     print(row[2].value)

# 대상으로 columns 단위로 Tuple 로 Wrap
#print(tuple(ws.columns))
# for column in tuple(ws.columns) : # for row in ws.iter_cols(min_row=1,max_row=5, min_col=3)
#     print(column[0].value)


#wb.save("_Sample_RPA.xlsx")

wb.close()
