from openpyxl import Workbook
from openpyxl import load_workbook # 파일 불러오기
from random import *

wb = Workbook() # New 워크북 생성
ws = wb.active # 현재 활성화된 Sheet 활성화

ws.title ="Cell_Create"
ws.sheet_properties.tabColor ="ff0066" # RGB 형태 

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["A4"] = 4
ws["A5"] = 5 


ws["B1"] = 1
ws["B2"] = 2
ws["B3"] = 3
ws["B4"] = 4
ws["B5"] = 5 

c=ws.cell(column=3,row=1, value=10)


print(ws["A1"].value, ws["A9"].value)
#or
print(ws.cell(row=1, column=1).value) # === A1 Cell과 같음

#print(c)

for xLoop in range(1,11):
    for yLoop in range(1,11):
        ws.cell(row=xLoop, column=yLoop, value=randint(0,100))


wb.save("_Sample_RPA.xlsx")


wb = load_workbook("_Sample_RPA.xlsx")
ws = wb.active 

# Row 와 Cloumn 이 정형화 된 Case
# for xLoop in range(1,11):
#     for yLoop in range(1,11):
#         print(ws.cell(row=xLoop, column=yLoop).value, end="")
#     print()


# Row 와 Cloumn 비정영화된 Case 
for xLoop in range(1,ws.max_row + 1):
    for yLoop in range(1,ws.max_column + 1):
        print(ws.cell(row=xLoop, column=yLoop).value, end="")
    print()



wb.close()
